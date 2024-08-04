#Projeto: automatização 2
#Fazer um programa que pegue dados de uma planilha e jogue dentro de um site.

def iniciar(valorPl, valorFol, valorLi):
    from selenium import webdriver # importa o web driver do selenium
    from webdriver_manager.chrome import ChromeDriverManager # importa o manager que vai gerenciar o webdriver e ver qual a versão mais recente do navegador do Chrome
    from selenium.webdriver.chrome.service import Service # importa o service para para instalar a versão mais recente
    import openpyxl
    from os import path
    from time import sleep

    def EnviarValor(num, valor):
        '''
        Essa função serve para enviar os valores da planilha para o site, recebendo um valor númerico da váriavel "c" para acessar uma
        posição do vetor "xpaths".
        '''
        try:
            # aqui verifica se o navegador está na última página para cadastrar um outro produto
            nav.find_element('xpath', f'{xpaths[num]}').send_keys(valor) # pede ao navegador para encontrar um xpath especifico para entregar um valor da planilha
            sleep(0.5)
        except Exception as e:
            print(f'Erro ao enviar valor: {e}')

    planilha = openpyxl.load_workbook(valorPl)  # pega valor digitado dentro da interface e procura no computador

    folha = planilha[valorFol] # acessa a folha dentro da planilha
    c = 0 # um contador para acessar a posição do xpath
    linha = 2
    qtd = int(valorLi) # quantas linhas devem ser automatizadas


    xpaths = ['//*[@id="productName"]', '//*[@id="productDescription"]', '//*[@id="productPrice"]', '//*[@id="productExpiry"]', 
            '//*[@id="productCode"]', '//*[@id="productColor"]', '//*[@id="productSize"]', '//*[@id="productMaterial"]', 
            '//*[@id="productManufacturer"]', '//*[@id="productOrigin"]', '/html/body/div/div/div[11]/input', '/html/body/div/div/div[12]/input',
            '/html/body/div/div/div[13]/input'] # é uma lista com todos os xpaths necessários para a automatização

    servico = Service(ChromeDriverManager().install()) # instala a versão mais recente do webdriver
    nav = webdriver.Chrome(service=servico)
    nav.get('https://fernandeslr.github.io/deploy/index.html') # acessa o site



    # Essas condicionais verificam e ajustam para o número correto já que o sistema sempre registra 1 a menos
    if qtd == 1:
        qtd = 2 # O sistema só registra o primeiro produto e não 2 produtos
    elif qtd >= 2:
        qtd += 1 # faz com que o sistema leia a quantidade exata de linhas colocada pelo usuário

    while linha <= qtd:
        c = 0
        for coluna in range(1, folha.max_column + 1): # vai passando de coluna em coluna da planilha
            celula = folha.cell(row=linha, column=coluna) # acessa a célula com o valor
            EnviarValor(c, celula.value) # passa o contador para a função
            c +=1
            if c == len(xpaths):
                nav.find_element('xpath', '/html/body/div/div/a').click()
            elif nav.current_url == 'https://fernandeslr.github.io/deploy/page2.html':
                nav.find_element('xpath', '/html/body/div/a').click()

        linha += 1 # passa para próxima linha
    print('\033[92mProdutos cadastrados com sucesso!\033[0m')
    return True

