import os
from openpyxl import load_workbook
def planilha(nome_arquivo):
    # Obtém o diretório atual onde o script está rodando
    diretorio_atual = os.getcwd()
    
    # Monta o caminho completo do arquivo
    caminho_completo = os.path.join(diretorio_atual, nome_arquivo)
    
    # Verifica se o arquivo existe
    if os.path.isfile(caminho_completo) and '.xlsx' in nome_arquivo:
        return True
    else:
        return False
    
def folha(nome_arquivo, nome_folha):
    # Verifica se a função da planilha retornou verdadeiro
    if planilha(nome_arquivo):
        try:
            # Carrega a planilha
            planilha_nome = load_workbook(nome_arquivo)
            
            # Verifica se o valor da folha existe na planilha
            if nome_folha in planilha_nome.sheetnames:
                return True
            
        except FileNotFoundError:
            return False
        
    # retorna falso se a planilha não existir
    return False

def linhas(qtd):
    # evita que o usuário coloque um valor negativo
    try:
        qtd = int(qtd)
        if qtd > 0:
            return True
        else:
            raise ValueError
    except ValueError:
        return False
    